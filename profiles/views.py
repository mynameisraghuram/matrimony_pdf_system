import json
import os

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Count, Max, Q, Subquery, OuterRef
from django.utils import timezone
from django.http import FileResponse, Http404, HttpResponse, HttpResponseNotAllowed
from django.shortcuts import get_object_or_404, redirect, render

import uuid

from profiles.forms import ProfileForm
from profiles.models import ActivityLog, GeneratedPDF, Interaction, Profile, ProfileNote
from profiles.utils.helpers import generate_profile_pdf
from profiles.services.sync_service import sync_profiles_from_sheet


@login_required
def activity_log(request):
    logs = ActivityLog.objects.select_related("profile").all()[:100]
    return render(request, "profiles/activity_log.html", {"logs": logs})


@login_required
def add_profile(request):
    duplicates = None
    if request.method == "POST":
        form = ProfileForm(request.POST)
        if form.is_valid():
            # Check for duplicates before saving
            first_name = form.cleaned_data.get("first_name", "").strip()
            last_name = form.cleaned_data.get("last_name", "").strip()
            contact = form.cleaned_data.get("contact_number", "").strip()

            dup_q = Q()
            if first_name and last_name:
                dup_q |= Q(first_name__iexact=first_name, last_name__iexact=last_name)
            if contact:
                dup_q |= Q(contact_number=contact)

            if dup_q and "confirm_save" not in request.POST:
                duplicates = Profile.objects.filter(dup_q)
                if duplicates.exists():
                    return render(request, "profiles/profile_form.html", {
                        "form": form, "title": "Add New Profile", "duplicates": duplicates,
                    })

            profile = form.save(commit=False)
            profile.profile_id = f"MANUAL-{uuid.uuid4().hex[:8]}"
            if profile.first_name and profile.last_name:
                profile.full_name = f"{profile.first_name} {profile.last_name}"
            elif profile.first_name:
                profile.full_name = profile.first_name
            profile.save()
            ActivityLog.objects.create(
                profile=profile, action="profile_synced",
                detail="Manually added",
            )
            messages.success(request, f"Profile {profile.display_id} created successfully.")
            return redirect("profiles:profile_detail", pk=profile.pk)
    else:
        form = ProfileForm()
    return render(request, "profiles/profile_form.html", {"form": form, "title": "Add New Profile"})


@login_required
def edit_profile(request, pk):
    profile = get_object_or_404(Profile, pk=pk)
    if request.method == "POST":
        form = ProfileForm(request.POST, instance=profile)
        if form.is_valid():
            profile = form.save(commit=False)
            if profile.first_name and profile.last_name:
                profile.full_name = f"{profile.first_name} {profile.last_name}"
            elif profile.first_name:
                profile.full_name = profile.first_name
            profile.save()
            messages.success(request, "Profile updated successfully.")
            return redirect("profiles:profile_detail", pk=pk)
    else:
        form = ProfileForm(instance=profile)
    return render(request, "profiles/profile_form.html", {
        "form": form, "title": f"Edit {profile.display_id}", "profile": profile,
    })


@login_required
def delete_profile(request, pk):
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])
    profile = get_object_or_404(Profile, pk=pk)
    display_id = profile.display_id
    # Delete associated photo file
    if profile.photo:
        photo_path = os.path.join(settings.BASE_DIR, profile.photo.path)
        if os.path.exists(photo_path):
            os.remove(photo_path)
    profile.delete()
    messages.success(request, f"Profile {display_id} deleted.")
    return redirect("profiles:profile_list")


@login_required
def add_note(request, pk):
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])
    profile = get_object_or_404(Profile, pk=pk)
    text = request.POST.get("note_text", "").strip()
    if text:
        ProfileNote.objects.create(profile=profile, text=text)
        messages.success(request, "Note added.")
    else:
        messages.warning(request, "Note cannot be empty.")
    return redirect("profiles:profile_detail", pk=pk)


@login_required
def delete_note(request, pk, note_id):
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])
    note = get_object_or_404(ProfileNote, id=note_id, profile__pk=pk)
    note.delete()
    messages.success(request, "Note deleted.")
    return redirect("profiles:profile_detail", pk=pk)


@login_required
def profile_list(request):
    profiles = Profile.objects.all()

    # Search
    q = request.GET.get("q", "").strip()
    if q:
        profiles = profiles.filter(
            Q(full_name__icontains=q)
            | Q(first_name__icontains=q)
            | Q(last_name__icontains=q)
            | Q(display_id__icontains=q)
            | Q(contact_number__icontains=q)
            | Q(email__icontains=q)
        )

    # Filter by looking_for (Bride/Groom)
    looking_for = request.GET.get("looking_for", "").strip()
    if looking_for:
        profiles = profiles.filter(looking_for__iexact=looking_for)

    # Filter by sub_caste
    sub_caste = request.GET.get("sub_caste", "").strip()
    if sub_caste:
        profiles = profiles.filter(sub_caste__iexact=sub_caste)

    # Filter by marital_status
    marital_status = request.GET.get("marital_status", "").strip()
    if marital_status:
        profiles = profiles.filter(marital_status__iexact=marital_status)

    # Filter by status
    status = request.GET.get("status", "").strip()
    if status:
        profiles = profiles.filter(status=status)

    # Sorting
    allowed_sort = {
        "display_id": "display_id",
        "full_name": "full_name",
        "looking_for": "looking_for",
        "sub_caste": "sub_caste",
        "contact_number": "contact_number",
        "status": "status",
        "updated_at": "updated_at",
    }
    sort = request.GET.get("sort", "").strip()
    order = request.GET.get("order", "").strip()
    if sort in allowed_sort:
        order_field = allowed_sort[sort]
        if order == "desc":
            order_field = f"-{order_field}"
        profiles = profiles.order_by(order_field)
    else:
        sort = ""
        profiles = profiles.order_by("-updated_at")

    # Annotate with last interaction date
    profiles = profiles.annotate(
        last_interaction=Max("interactions__created_at"),
    )

    # Pagination
    paginator = Paginator(profiles, 25)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # Get distinct values for filter dropdowns
    all_profiles = Profile.objects.all()
    sub_castes = (
        all_profiles.exclude(sub_caste__isnull=True)
        .exclude(sub_caste="")
        .values_list("sub_caste", flat=True)
        .distinct()
        .order_by("sub_caste")
    )
    marital_statuses = (
        all_profiles.exclude(marital_status__isnull=True)
        .exclude(marital_status="")
        .values_list("marital_status", flat=True)
        .distinct()
        .order_by("marital_status")
    )

    # Dashboard stats
    total_profiles = all_profiles.count()
    brides = all_profiles.filter(looking_for__iexact="bride").count()
    grooms = all_profiles.filter(looking_for__iexact="groom").count()
    with_photo = all_profiles.exclude(photo="").exclude(photo__isnull=True).count()
    total_pdfs = GeneratedPDF.objects.count()
    week_ago = timezone.now() - timezone.timedelta(days=7)
    pdfs_this_week = GeneratedPDF.objects.filter(generated_at__gte=week_ago).count()
    from datetime import date as _date
    overdue_followups = Interaction.objects.filter(
        follow_up_date__isnull=False, follow_up_date__lt=_date.today()
    ).count()
    today_followups = Interaction.objects.filter(follow_up_date=_date.today()).count()

    # Chart data — profiles by sub caste
    sub_caste_data = list(
        all_profiles.exclude(sub_caste__isnull=True)
        .exclude(sub_caste="")
        .values("sub_caste")
        .annotate(count=Count("id"))
        .order_by("-count")[:10]
    )

    # Chart data — profiles by status
    status_data = list(
        all_profiles.values("status")
        .annotate(count=Count("id"))
        .order_by("status")
    )

    context = {
        "profiles": page_obj,
        "page_obj": page_obj,
        "q": q,
        "looking_for": looking_for,
        "sub_caste": sub_caste,
        "marital_status": marital_status,
        "status": status,
        "sort": sort,
        "order": order,
        "sub_castes": sub_castes,
        "marital_statuses": marital_statuses,
        "status_choices": Profile.STATUS_CHOICES,
        "stats": {
            "total": total_profiles,
            "brides": brides,
            "grooms": grooms,
            "with_photo": with_photo,
            "total_pdfs": total_pdfs,
            "pdfs_this_week": pdfs_this_week,
            "overdue_followups": overdue_followups,
            "today_followups": today_followups,
        },
        "chart_looking_for": json.dumps({"Grooms": grooms, "Brides": brides}),
        "chart_sub_caste": json.dumps(
            {d["sub_caste"]: d["count"] for d in sub_caste_data}
        ),
        "chart_status": json.dumps(
            {d["status"].replace("_", " ").title(): d["count"] for d in status_data}
        ),
    }
    return render(request, "profiles/profile_list.html", context)


@login_required
def sync_sheet(request):
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])

    try:
        result = sync_profiles_from_sheet()
        parts = []
        if result["created"]:
            parts.append(f"{result['created']} new")
        if result["updated"]:
            parts.append(f"{result['updated']} updated")
        if result["invalid"]:
            parts.append(f"{result['invalid']} invalid")
        if result["errors"]:
            parts.append(f"{result['errors']} errors")

        summary = ", ".join(parts) if parts else "no rows found"
        ActivityLog.objects.create(
            action="profile_synced",
            detail=f"{result['total']} rows fetched: {summary}",
        )
        messages.success(request, f"Sync complete — {result['total']} rows fetched: {summary}.")
    except Exception as e:
        messages.error(request, f"Sync failed: {str(e)}")

    return redirect("profiles:profile_list")


@login_required
def profile_detail(request, pk):
    profile = get_object_or_404(Profile, pk=pk)
    pdfs = profile.pdfs.all().order_by("-generated_at")
    recent_activity = profile.activities.all()[:10]
    notes = profile.notes.all()
    interactions = profile.interactions.select_related("logged_by").all()[:20]
    from datetime import date
    return render(request, "profiles/profile_detail.html", {
        "profile": profile, "pdfs": pdfs, "recent_activity": recent_activity,
        "notes": notes, "interactions": interactions,
        "interaction_types": Interaction.TYPE_CHOICES,
        "outcome_choices": Interaction.OUTCOME_CHOICES,
        "today": date.today(),
    })


@login_required
def generate_pdf(request, pk, tier="premium"):
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])

    if tier not in ("standard", "premium"):
        tier = "premium"

    profile = get_object_or_404(Profile, pk=pk)
    tier_label = tier.capitalize()

    try:
        file_path = generate_profile_pdf(profile, tier=tier)
        version = profile.pdfs.filter(tier=tier).count() + 1
        GeneratedPDF.objects.create(
            profile=profile,
            file_path=file_path,
            version=version,
            tier=tier,
            template_used=tier,
        )
        ActivityLog.objects.create(
            profile=profile,
            action="pdf_generated",
            detail=f"{tier_label} PDF v{version}",
        )
        messages.success(request, f"{tier_label} PDF generated successfully (v{version}).")
    except Exception as e:
        messages.error(request, f"{tier_label} PDF generation failed: {str(e)}")

    return redirect("profiles:profile_detail", pk=pk)


@login_required
def bulk_generate_pdf(request):
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])

    tier = request.POST.get("tier", "standard")
    if tier not in ("standard", "premium"):
        tier = "standard"

    profile_ids = request.POST.getlist("profile_ids")
    if not profile_ids:
        messages.warning(request, "No profiles selected.")
        return redirect("profiles:profile_list")

    profiles = Profile.objects.filter(pk__in=profile_ids)
    success_count = 0
    error_count = 0
    tier_label = tier.capitalize()

    for profile in profiles:
        try:
            file_path = generate_profile_pdf(profile, tier=tier)
            version = profile.pdfs.filter(tier=tier).count() + 1
            GeneratedPDF.objects.create(
                profile=profile,
                file_path=file_path,
                version=version,
                tier=tier,
                template_used=tier,
            )
            success_count += 1
        except Exception:
            error_count += 1

    parts = []
    if success_count:
        parts.append(f"{success_count} {tier_label} PDFs generated")
    if error_count:
        parts.append(f"{error_count} failed")
    ActivityLog.objects.create(
        action="bulk_pdf_generated",
        detail=f"{', '.join(parts)}",
    )
    messages.success(request, f"Bulk generation complete — {', '.join(parts)}.")
    return redirect("profiles:profile_list")


@login_required
def upload_photo(request, pk):
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])

    profile = get_object_or_404(Profile, pk=pk)

    if "photo" not in request.FILES:
        messages.error(request, "No photo file selected.")
        return redirect("profiles:profile_detail", pk=pk)

    photo = request.FILES["photo"]

    # Validate file type
    allowed = (".jpg", ".jpeg", ".png", ".webp")
    if not photo.name.lower().endswith(allowed):
        messages.error(request, "Only JPG, PNG, and WebP images are allowed.")
        return redirect("profiles:profile_detail", pk=pk)

    # Delete old photo if exists
    if profile.photo:
        old_path = os.path.join(settings.BASE_DIR, profile.photo.path)
        if os.path.exists(old_path):
            os.remove(old_path)

    # Resize and compress the photo
    from PIL import Image
    from io import BytesIO
    from django.core.files.uploadedfile import InMemoryUploadedFile

    img = Image.open(photo)
    if img.mode in ("RGBA", "P"):
        img = img.convert("RGB")

    max_size = (400, 400)
    img.thumbnail(max_size, Image.LANCZOS)

    buffer = BytesIO()
    img.save(buffer, format="JPEG", quality=85, optimize=True)
    buffer.seek(0)

    # Build filename as .jpg
    base_name = os.path.splitext(photo.name)[0]
    resized_file = InMemoryUploadedFile(
        buffer, "photo", f"{base_name}.jpg", "image/jpeg", buffer.getbuffer().nbytes, None
    )

    profile.photo = resized_file
    profile.save()
    ActivityLog.objects.create(
        profile=profile, action="photo_uploaded",
        detail=f"{photo.name} (resized to {img.size[0]}x{img.size[1]})",
    )
    messages.success(request, f"Photo uploaded and resized to {img.size[0]}x{img.size[1]}.")
    return redirect("profiles:profile_detail", pk=pk)


@login_required
def delete_photo(request, pk):
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])

    profile = get_object_or_404(Profile, pk=pk)

    if profile.photo:
        old_path = os.path.join(settings.BASE_DIR, profile.photo.path)
        if os.path.exists(old_path):
            os.remove(old_path)
        profile.photo = None
        profile.save()
        ActivityLog.objects.create(
            profile=profile, action="photo_deleted",
        )
        messages.success(request, "Photo removed.")
    return redirect("profiles:profile_detail", pk=pk)


@login_required
def preview_pdf(request, pk, pdf_id):
    pdf = get_object_or_404(GeneratedPDF, id=pdf_id, profile__pk=pk)
    abs_path = os.path.join(settings.BASE_DIR, pdf.file_path)

    if not os.path.exists(abs_path):
        raise Http404("PDF file not found on disk.")

    ActivityLog.objects.create(
        profile=pdf.profile, action="pdf_previewed",
        detail=f"{pdf.tier.capitalize()} v{pdf.version}",
    )
    return FileResponse(
        open(abs_path, "rb"),
        content_type="application/pdf",
        filename=os.path.basename(abs_path),
    )


@login_required
def download_pdf(request, pk, pdf_id):
    pdf = get_object_or_404(GeneratedPDF, id=pdf_id, profile__pk=pk)
    abs_path = os.path.join(settings.BASE_DIR, pdf.file_path)

    if not os.path.exists(abs_path):
        raise Http404("PDF file not found on disk.")

    ActivityLog.objects.create(
        profile=pdf.profile, action="pdf_downloaded",
        detail=f"{pdf.tier.capitalize()} v{pdf.version}",
    )
    return FileResponse(
        open(abs_path, "rb"),
        as_attachment=True,
        filename=os.path.basename(abs_path),
    )


@login_required
def export_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    profiles = Profile.objects.all()

    # Apply same filters as profile_list
    q = request.GET.get("q", "").strip()
    if q:
        profiles = profiles.filter(
            Q(full_name__icontains=q)
            | Q(first_name__icontains=q)
            | Q(last_name__icontains=q)
            | Q(display_id__icontains=q)
            | Q(contact_number__icontains=q)
            | Q(email__icontains=q)
        )
    looking_for = request.GET.get("looking_for", "").strip()
    if looking_for:
        profiles = profiles.filter(looking_for__iexact=looking_for)
    sub_caste = request.GET.get("sub_caste", "").strip()
    if sub_caste:
        profiles = profiles.filter(sub_caste__iexact=sub_caste)
    marital_status = request.GET.get("marital_status", "").strip()
    if marital_status:
        profiles = profiles.filter(marital_status__iexact=marital_status)
    status = request.GET.get("status", "").strip()
    if status:
        profiles = profiles.filter(status=status)

    profiles = profiles.order_by("-updated_at")

    wb = Workbook()
    ws = wb.active
    ws.title = "Profiles"

    columns = [
        ("Profile ID", "display_id", 12),
        ("Full Name", "full_name", 25),
        ("Status", "status", 12),
        ("Looking For", "looking_for", 12),
        ("Date of Birth", "date_of_birth", 14),
        ("Height", "height", 10),
        ("Marital Status", "marital_status", 14),
        ("Sub Caste", "sub_caste", 14),
        ("Gothram", "gothram", 12),
        ("Star", "star", 12),
        ("Rasi", "rasi", 12),
        ("Graduation", "graduation", 20),
        ("Masters", "masters", 20),
        ("Designation", "designation", 20),
        ("Company", "company_name", 20),
        ("Salary", "salary", 12),
        ("Job Location", "job_location", 15),
        ("Father Name", "father_name", 20),
        ("Father Occupation", "father_occupation", 20),
        ("Mother Name", "mother_name", 20),
        ("Mother Occupation", "mother_occupation", 20),
        ("Siblings", "siblings", 25),
        ("Contact", "contact_number", 15),
        ("Email", "email", 25),
    ]

    # Header row
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="C42027", end_color="C42027", fill_type="solid")
    for col_idx, (label, _, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = width

    # Data rows
    for row_idx, profile in enumerate(profiles, 2):
        for col_idx, (_, field, _) in enumerate(columns, 1):
            ws.cell(row=row_idx, column=col_idx, value=getattr(profile, field, "") or "")

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="profiles_export.xlsx"'
    wb.save(response)
    return response


@login_required
def email_pdf(request, pk, pdf_id):
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])

    from django.core.mail import EmailMessage

    profile = get_object_or_404(Profile, pk=pk)
    pdf = get_object_or_404(GeneratedPDF, id=pdf_id, profile=profile)
    abs_path = os.path.join(settings.BASE_DIR, pdf.file_path)

    if not os.path.exists(abs_path):
        messages.error(request, "PDF file not found on disk.")
        return redirect("profiles:profile_detail", pk=pk)

    recipient = request.POST.get("email", "").strip()
    if not recipient:
        messages.error(request, "No email address provided.")
        return redirect("profiles:profile_detail", pk=pk)

    # Check email settings are configured
    if not settings.EMAIL_HOST_USER:
        messages.error(request, "Email is not configured. Set EMAIL_HOST_USER and EMAIL_HOST_PASSWORD in .env")
        return redirect("profiles:profile_detail", pk=pk)

    company = os.environ.get("COMPANY_NAME", "Reddy & Reddy Matrimony")
    tier_label = pdf.tier.capitalize()
    name = profile.full_name or profile.display_id

    subject = f"{tier_label} Profile - {name} ({profile.display_id}) | {company}"
    body = (
        f"Dear Sir/Madam,\n\n"
        f"Please find attached the {tier_label} Profile for:\n\n"
        f"  Name: {name}\n"
        f"  Profile ID: {profile.display_id}\n"
    )
    if profile.looking_for:
        body += f"  Looking For: {profile.looking_for}\n"
    if profile.sub_caste:
        body += f"  Sub Caste: {profile.sub_caste}\n"
    if profile.height:
        body += f"  Height: {profile.height}\n"

    body += (
        f"\nFor more details, please contact us:\n"
        f"  Phone: +91 9368111222\n"
        f"  WhatsApp: wa.me/9368111222\n\n"
        f"Warm regards,\n{company}\n"
    )

    try:
        email = EmailMessage(
            subject=subject,
            body=body,
            to=[recipient],
        )
        email.attach_file(abs_path)
        email.send()
        ActivityLog.objects.create(
            profile=profile, action="pdf_emailed",
            detail=f"{tier_label} v{pdf.version} to {recipient}",
        )
        messages.success(request, f"PDF emailed to {recipient} successfully.")
    except Exception as e:
        messages.error(request, f"Email failed: {str(e)}")

    return redirect("profiles:profile_detail", pk=pk)


@login_required
def import_profiles(request):
    import csv
    from io import TextIOWrapper
    from openpyxl import load_workbook

    # Column header -> model field mapping (case-insensitive)
    HEADER_MAP = {
        "first name": "first_name",
        "last name": "last_name",
        "full name": "full_name",
        "looking for": "looking_for",
        "marital status": "marital_status",
        "date of birth": "date_of_birth",
        "time of birth": "time_of_birth",
        "place of birth": "place_of_birth",
        "height": "height",
        "star": "star",
        "rasi": "rasi",
        "sub caste": "sub_caste",
        "gothram": "gothram",
        "schooling": "schooling",
        "graduation": "graduation",
        "masters": "masters",
        "designation": "designation",
        "company": "company_name",
        "company name": "company_name",
        "salary": "salary",
        "years of exp": "years_of_exp",
        "experience": "years_of_exp",
        "job location": "job_location",
        "visa status": "visa_status",
        "father name": "father_name",
        "father occupation": "father_occupation",
        "father native": "father_native",
        "mother name": "mother_name",
        "mother occupation": "mother_occupation",
        "mother native": "mother_native",
        "siblings": "siblings",
        "parents staying": "parents_staying",
        "actual property": "actual_property",
        "shared property": "shared_property",
        "expected property": "expected_property",
        "preferred height": "preferred_height",
        "age gap": "age_gap",
        "preferred sub caste": "preferred_sub_caste",
        "astrology": "astrology",
        "looking country": "looking_country",
        "looking state": "looking_state",
        "education preference": "education_preference",
        "career preferences": "career_preferences",
        "special conditions": "special_conditions",
        "contact": "contact_number",
        "contact number": "contact_number",
        "phone": "contact_number",
        "second contact": "second_contact_number",
        "second contact number": "second_contact_number",
        "email": "email",
        "status": "status",
    }

    if request.method == "POST" and request.FILES.get("file"):
        uploaded = request.FILES["file"]
        fname = uploaded.name.lower()

        rows = []
        headers = []

        try:
            if fname.endswith(".csv"):
                text_file = TextIOWrapper(uploaded.file, encoding="utf-8-sig")
                reader = csv.reader(text_file)
                raw_headers = next(reader)
                headers = [h.strip() for h in raw_headers]
                for row in reader:
                    if any(cell.strip() for cell in row):
                        rows.append(row)
            elif fname.endswith((".xlsx", ".xls")):
                wb = load_workbook(uploaded, read_only=True, data_only=True)
                ws = wb.active
                all_rows = list(ws.iter_rows(values_only=True))
                if not all_rows:
                    messages.error(request, "The file appears to be empty.")
                    return redirect("profiles:import_profiles")
                headers = [str(h).strip() if h else "" for h in all_rows[0]]
                for row in all_rows[1:]:
                    str_row = [str(cell).strip() if cell is not None else "" for cell in row]
                    if any(str_row):
                        rows.append(str_row)
                wb.close()
            else:
                messages.error(request, "Unsupported file format. Please upload .csv or .xlsx files.")
                return redirect("profiles:import_profiles")
        except Exception as e:
            messages.error(request, f"Error reading file: {e}")
            return redirect("profiles:import_profiles")

        if not rows:
            messages.error(request, "No data rows found in the file.")
            return redirect("profiles:import_profiles")

        # Map headers to model fields
        field_map = {}  # column index -> model field name
        for idx, header in enumerate(headers):
            key = header.lower().strip()
            if key in HEADER_MAP:
                field_map[idx] = HEADER_MAP[key]

        if not field_map:
            messages.error(request, "No recognized column headers found. Expected headers like: Full Name, Contact Number, Sub Caste, etc.")
            return redirect("profiles:import_profiles")

        # Import rows
        created = 0
        skipped = 0
        errors = []

        for row_num, row in enumerate(rows, start=2):
            data = {}
            for col_idx, field_name in field_map.items():
                if col_idx < len(row):
                    val = row[col_idx].strip() if isinstance(row[col_idx], str) else str(row[col_idx]).strip()
                    if val and val.lower() != "none":
                        data[field_name] = val

            if not data:
                skipped += 1
                continue

            # Build full_name if not provided
            if "full_name" not in data and ("first_name" in data or "last_name" in data):
                parts = [data.get("first_name", ""), data.get("last_name", "")]
                data["full_name"] = " ".join(p for p in parts if p)

            # Check for duplicates by contact_number or full_name
            dup_q = Q()
            if data.get("contact_number"):
                dup_q |= Q(contact_number=data["contact_number"])
            if data.get("full_name") and data.get("full_name") != "":
                dup_q |= Q(full_name__iexact=data["full_name"])

            if dup_q and Profile.objects.filter(dup_q).exists():
                skipped += 1
                continue

            try:
                # Validate status
                if "status" in data:
                    valid_statuses = [s[0] for s in Profile.STATUS_CHOICES]
                    if data["status"].lower() not in valid_statuses:
                        data["status"] = "active"
                    else:
                        data["status"] = data["status"].lower()

                profile = Profile(
                    profile_id=f"IMPORT-{uuid.uuid4().hex[:8]}",
                    **data,
                )
                profile.save()
                created += 1
            except Exception as e:
                errors.append(f"Row {row_num}: {e}")

        if created:
            ActivityLog.objects.create(
                action="profile_synced",
                detail=f"Imported {created} profiles from {uploaded.name}",
            )
            messages.success(request, f"Successfully imported {created} profile(s).")
        if skipped:
            messages.info(request, f"Skipped {skipped} row(s) (duplicate or empty).")
        if errors:
            messages.warning(request, f"{len(errors)} error(s): {'; '.join(errors[:5])}")

        return redirect("profiles:profile_list")

    # GET - show upload form
    return render(request, "profiles/import_profiles.html")


@login_required
def compare_profiles(request):
    ids = request.GET.getlist("ids")
    if len(ids) < 2:
        messages.warning(request, "Please select at least 2 profiles to compare.")
        return redirect("profiles:profile_list")
    if len(ids) > 4:
        ids = ids[:4]

    profiles = Profile.objects.filter(pk__in=ids)
    if profiles.count() < 2:
        messages.error(request, "Selected profiles not found.")
        return redirect("profiles:profile_list")

    # Define comparison fields grouped by section
    sections = [
        ("Basic Details", [
            ("Full Name", "full_name"),
            ("Date of Birth", "date_of_birth"),
            ("Time of Birth", "time_of_birth"),
            ("Place of Birth", "place_of_birth"),
            ("Height", "height"),
            ("Looking For", "looking_for"),
            ("Marital Status", "marital_status"),
            ("Star", "star"),
            ("Rasi", "rasi"),
            ("Sub Caste", "sub_caste"),
            ("Gothram", "gothram"),
        ]),
        ("Education & Career", [
            ("Schooling", "schooling"),
            ("Graduation", "graduation"),
            ("Masters", "masters"),
            ("Designation", "designation"),
            ("Company", "company_name"),
            ("Salary", "salary"),
            ("Experience", "years_of_exp"),
            ("Job Location", "job_location"),
            ("Visa Status", "visa_status"),
        ]),
        ("Family Details", [
            ("Father Name", "father_name"),
            ("Father Occupation", "father_occupation"),
            ("Father Native", "father_native"),
            ("Mother Name", "mother_name"),
            ("Mother Occupation", "mother_occupation"),
            ("Mother Native", "mother_native"),
            ("Siblings", "siblings"),
            ("Parents Staying", "parents_staying"),
        ]),
        ("Property", [
            ("Actual Property", "actual_property"),
            ("Shared Property", "shared_property"),
            ("Expected Property", "expected_property"),
        ]),
        ("Partner Preferences", [
            ("Preferred Height", "preferred_height"),
            ("Age Gap", "age_gap"),
            ("Preferred Sub Caste", "preferred_sub_caste"),
            ("Astrology", "astrology"),
            ("Looking Country", "looking_country"),
            ("Looking State", "looking_state"),
            ("Education Preference", "education_preference"),
            ("Career Preferences", "career_preferences"),
            ("Special Conditions", "special_conditions"),
        ]),
        ("Contact", [
            ("Email", "email"),
            ("Contact Number", "contact_number"),
            ("Second Contact", "second_contact_number"),
        ]),
    ]

    return render(request, "profiles/compare_profiles.html", {
        "profiles": profiles,
        "sections": sections,
    })


@login_required
def add_interaction(request, pk):
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])

    profile = get_object_or_404(Profile, pk=pk)

    interaction_type = request.POST.get("interaction_type", "phone_call")
    outcome = request.POST.get("outcome", "info_gathered")
    summary = request.POST.get("summary", "").strip()
    follow_up = request.POST.get("follow_up_date", "").strip()
    audio = request.FILES.get("audio_file")

    # If audio file uploaded, transcribe and summarize
    transcript = ""
    if audio:
        allowed_ext = (".mp3", ".wav", ".m4a", ".ogg", ".webm", ".mp4")
        if not audio.name.lower().endswith(allowed_ext):
            messages.error(request, f"Unsupported audio format. Use: {', '.join(allowed_ext)}")
            return redirect("profiles:profile_detail", pk=pk)

        # Save audio temporarily for Whisper API
        import tempfile
        ext = os.path.splitext(audio.name)[1]
        with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
            for chunk in audio.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name

        try:
            from profiles.services.call_transcriber import transcribe_audio, summarize_transcript
            messages.info(request, "Transcribing audio... this may take a moment.")
            transcript = transcribe_audio(tmp_path)
            if not summary:
                profile_name = profile.full_name or profile.display_id
                summary = summarize_transcript(transcript, profile_name)
        except Exception as e:
            messages.error(request, f"Audio transcription failed: {e}")
            if not summary:
                summary = "(Audio uploaded but transcription failed)"
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)

    if not summary:
        messages.warning(request, "Please provide a summary or upload an audio file.")
        return redirect("profiles:profile_detail", pk=pk)

    interaction = Interaction(
        profile=profile,
        interaction_type=interaction_type,
        outcome=outcome,
        summary=summary,
        transcript=transcript,
        logged_by=request.user,
    )

    if audio:
        interaction.audio_file = audio

    if follow_up:
        try:
            from datetime import date
            interaction.follow_up_date = date.fromisoformat(follow_up)
        except ValueError:
            pass

    interaction.save()
    messages.success(request, f"Interaction logged: {interaction.get_interaction_type_display()}")
    return redirect("profiles:profile_detail", pk=pk)


@login_required
def delete_interaction(request, pk, interaction_id):
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])
    interaction = get_object_or_404(Interaction, id=interaction_id, profile__pk=pk)
    interaction.delete()
    messages.success(request, "Interaction deleted.")
    return redirect("profiles:profile_detail", pk=pk)


@login_required
def follow_ups(request):
    """Show all interactions with pending follow-up dates."""
    from datetime import date
    today = date.today()
    upcoming = Interaction.objects.filter(
        follow_up_date__isnull=False,
        follow_up_date__gte=today,
    ).select_related("profile", "logged_by").order_by("follow_up_date")[:50]

    overdue = Interaction.objects.filter(
        follow_up_date__isnull=False,
        follow_up_date__lt=today,
    ).select_related("profile", "logged_by").order_by("follow_up_date")[:50]

    return render(request, "profiles/follow_ups.html", {
        "upcoming": upcoming,
        "overdue": overdue,
        "today": today,
    })


@login_required
def backup_database(request):
    """Download the SQLite database as a backup file."""
    if not request.user.is_superuser:
        messages.error(request, "Only admin users can download backups.")
        return redirect("profiles:profile_list")

    import shutil
    import tempfile
    from datetime import datetime

    db_path = settings.DATABASES["default"]["NAME"]
    if not os.path.exists(db_path):
        messages.error(request, "Database file not found.")
        return redirect("profiles:profile_list")

    # Copy to temp file to avoid locking issues
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".sqlite3")
    tmp.close()
    shutil.copy2(db_path, tmp.name)

    response = FileResponse(
        open(tmp.name, "rb"),
        as_attachment=True,
        filename=f"matrimony_backup_{timestamp}.sqlite3",
    )
    # Clean up temp file after response is sent
    response._tmp_path = tmp.name
    return response


@login_required
def restore_database(request):
    """Restore the SQLite database from an uploaded backup file."""
    if not request.user.is_superuser:
        messages.error(request, "Only admin users can restore backups.")
        return redirect("profiles:profile_list")

    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])

    uploaded = request.FILES.get("backup_file")
    if not uploaded:
        messages.error(request, "No backup file selected.")
        return redirect("profiles:backup_page")

    if not uploaded.name.endswith(".sqlite3"):
        messages.error(request, "Invalid file. Please upload a .sqlite3 backup file.")
        return redirect("profiles:backup_page")

    import shutil
    import tempfile
    import sqlite3
    from datetime import datetime

    # Save uploaded file to temp location
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".sqlite3")
    for chunk in uploaded.chunks():
        tmp.write(chunk)
    tmp.close()

    # Validate it's a real SQLite database
    try:
        conn = sqlite3.connect(tmp.name)
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='profiles_profile'")
        if not cursor.fetchone():
            conn.close()
            os.remove(tmp.name)
            messages.error(request, "Invalid backup: does not contain the profiles table.")
            return redirect("profiles:backup_page")
        conn.close()
    except Exception as e:
        os.remove(tmp.name)
        messages.error(request, f"Invalid database file: {e}")
        return redirect("profiles:backup_page")

    db_path = str(settings.DATABASES["default"]["NAME"])

    # Create a backup of current DB before replacing
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_dir = os.path.join(settings.BASE_DIR, "backups")
    os.makedirs(backup_dir, exist_ok=True)
    pre_restore_backup = os.path.join(backup_dir, f"pre_restore_{timestamp}.sqlite3")
    shutil.copy2(db_path, pre_restore_backup)

    # Replace the database
    try:
        from django import db
        db.connections.close_all()
        shutil.copy2(tmp.name, db_path)
        messages.success(request, f"Database restored successfully from {uploaded.name}. Pre-restore backup saved.")
    except Exception as e:
        messages.error(request, f"Restore failed: {e}")
    finally:
        os.remove(tmp.name)

    return redirect("profiles:backup_page")


@login_required
def backup_page(request):
    """Show backup/restore page with list of local backups."""
    if not request.user.is_superuser:
        messages.error(request, "Only admin users can access backups.")
        return redirect("profiles:profile_list")

    backups = []
    backup_dir = os.path.join(settings.BASE_DIR, "backups")
    if os.path.exists(backup_dir):
        for fname in sorted(os.listdir(backup_dir), reverse=True):
            if fname.endswith(".sqlite3"):
                fpath = os.path.join(backup_dir, fname)
                size_mb = os.path.getsize(fpath) / (1024 * 1024)
                from datetime import datetime
                mtime = datetime.fromtimestamp(os.path.getmtime(fpath))
                backups.append({
                    "name": fname,
                    "size": f"{size_mb:.2f} MB",
                    "date": mtime,
                })

    db_path = settings.DATABASES["default"]["NAME"]
    db_size = os.path.getsize(db_path) / (1024 * 1024) if os.path.exists(db_path) else 0
    profile_count = Profile.objects.count()

    return render(request, "profiles/backup.html", {
        "backups": backups,
        "db_size": f"{db_size:.2f} MB",
        "profile_count": profile_count,
    })
